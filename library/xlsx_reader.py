################################  MODULE  INFO  ################################
# Author: David  Cobos
# Cisco Systems Solutions Integrations Architect
# Mail: cdcobos1999@gmail.com  / dacobos@cisco.com
################################  MODULE  INFO  ################################

# Read a xlsx file and return a dictionary with all the values

from openpyxl import Workbook
from openpyxl import load_workbook

def readxlsx(db_sap):
    wb = load_workbook(db_sap)
    ws = wb.active
    data = []
    for row in range (1,ws.max_row+1):
        values = []
        for col in ws.iter_cols(min_row=row, max_col=6, max_row=row):
                for cell in col:
                    values.append(cell.value)
        data.append(values)

    return data
